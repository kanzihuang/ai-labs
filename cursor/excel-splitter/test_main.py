#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from main import process_excel, load_config

def test_excel_sheet_copy():
    """Test that the result sheet has the same headers, data and formatting as source sheet."""
    # Test configuration
    test_config = {
        'input': {
            'path': 'test_input.xlsx',
            'sheet': {
                'source': {
                    'name': '工资',
                    'columns': {
                        'employee_id': '工号',
                        'project_id': '费用所属中心',
                        'project_category': '费用类别',
                        'project_hours': '实际出勤'
                    }
                }
            }
        },
        'output': {
            'path': 'test_output.xlsx',
            'sheet': {
                'result': {
                    'name': '工资拆分'
                }
            }
        }
    }

    # Create test input file
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    ws = wb.create_sheet('工资')
    
    # Add test data
    headers = ['工号', '费用所属中心', '费用类别', '实际出勤']
    ws.append(headers)
    test_data = [
        ['001', '中心A', '类别1', 8],
        ['002', '中心B', '类别2', 6]
    ]
    for row in test_data:
        ws.append(row)

    # Add some formatting
    bold_font = Font(bold=True)
    for cell in ws[1]:  # Format headers
        cell.font = bold_font
    
    # Save test input file
    wb.save('test_input.xlsx')

    try:
        # Process the Excel file
        process_excel(test_config)

        # Load both workbooks for comparison
        input_wb = load_workbook('test_input.xlsx')
        output_wb = load_workbook('test_output.xlsx')

        source_sheet = input_wb['工资']
        result_sheet = output_wb['工资拆分']

        # Test 1: Compare headers
        source_headers = [cell.value for cell in source_sheet[1]]
        result_headers = [cell.value for cell in result_sheet[1]]
        assert source_headers == result_headers, "Headers do not match"

        # Test 2: Compare data
        for row_idx in range(2, source_sheet.max_row + 1):
            source_row = [cell.value for cell in source_sheet[row_idx]]
            result_row = [cell.value for cell in result_sheet[row_idx]]
            assert source_row == result_row, f"Data in row {row_idx} does not match"

        # Test 3: Compare formatting
        for row in range(1, source_sheet.max_row + 1):
            for col in range(1, source_sheet.max_column + 1):
                source_cell = source_sheet.cell(row=row, column=col)
                result_cell = result_sheet.cell(row=row, column=col)
                
                # Compare font
                assert source_cell.font.bold == result_cell.font.bold, f"Font bold mismatch at {row},{col}"
                
                # Compare alignment
                assert source_cell.alignment.horizontal == result_cell.alignment.horizontal, f"Horizontal alignment mismatch at {row},{col}"
                assert source_cell.alignment.vertical == result_cell.alignment.vertical, f"Vertical alignment mismatch at {row},{col}"

        # Test 4: Compare column widths
        for col in source_sheet.column_dimensions:
            assert source_sheet.column_dimensions[col].width == result_sheet.column_dimensions[col].width, f"Column width mismatch for column {col}"

        # Test 5: Compare row heights
        for row in source_sheet.row_dimensions:
            assert source_sheet.row_dimensions[row].height == result_sheet.row_dimensions[row].height, f"Row height mismatch for row {row}"

    finally:
        # Clean up test files
        if os.path.exists('test_input.xlsx'):
            os.remove('test_input.xlsx')
        if os.path.exists('test_output.xlsx'):
            os.remove('test_output.xlsx') 