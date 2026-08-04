#!/usr/bin/env python
# -*- coding: utf-8 -*-

import argparse
import math
import os
import sys
import time
import traceback
import yaml
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils.exceptions import InvalidFileException
from copy import copy

def load_config(config_path):
    """Load configuration from YAML file."""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    except Exception as e:
        fatal(f"Error loading config file: {e}")

def copy_cell_style(source_cell, target_cell):
    """Copy cell style from source to target cell."""
    if source_cell.has_style and source_cell._style is not None:
        target_cell._style = source_cell._style

def get_column_index(headers, column_name):
    """Get column index (1-based) for a given column name."""
    try:
        return headers.index(column_name) + 1
    except ValueError:
        return None

import re


class FormulaError(Exception):
    """Raised when a formula is invalid or evaluation fails."""
    pass


def tokenize_formula(formula, known_names):
    """Tokenize a formula string into tokens (column names, operators, numbers, parens).

    Uses longest-match regex against known_names so column names containing
    parentheses (e.g. 税前应发工资总额(不含差补)) are matched as single tokens.
    """
    # Build alternation pattern: longest names first so they match before substrings
    if known_names:
        str_names = [str(n) for n in known_names]
        escaped = [re.escape(n) for n in sorted(str_names, key=len, reverse=True)]
        name_pattern = '|'.join(escaped)
        pattern = re.compile(r'(?:' + name_pattern + r')|(?:\d+(?:\.\d+)?)')
    else:
        pattern = re.compile(r'(?:\d+(?:\.\d+)?)')

    tokens = []
    pos = 0

    def scan_gap(text):
        """Scan a gap between regex matches, collecting operators and reporting unknown words."""
        i = 0
        while i < len(text):
            ch = text[i]
            if ch.isspace():
                i += 1
                continue
            if ch in '+-*/()':
                tokens.append(ch)
                i += 1
                continue
            # Collect consecutive non-operator chars as unknown word
            j = i
            while j < len(text) and not text[j].isspace() and text[j] not in '+-*/()':
                j += 1
            word = text[i:j]
            raise FormulaError(f"无法识别的列名 '{word}'")

    for m in pattern.finditer(formula):
        gap = formula[pos:m.start()]
        scan_gap(gap)
        tokens.append(m.group(0))
        pos = m.end()

    # Process trailing characters
    scan_gap(formula[pos:])

    return tokens


def parse_formula(tokens):
    """Recursive-descent parser for formula tokens.

    Grammar:
      expr   -> term (('+' | '-') term)*
      term   -> factor (('*' | '/') factor)*
      factor -> NAME | NUMBER | '(' expr ')' | '-' factor
    """
    pos = [0]  # mutable position tracker

    def peek():
        return tokens[pos[0]] if pos[0] < len(tokens) else None

    def consume():
        t = peek()
        if t is not None:
            pos[0] += 1
        return t

    def expr():
        left = term()
        while peek() in ('+', '-'):
            op = consume()
            right = term()
            left = (op, left, right)
        return left

    def term():
        left = factor()
        while peek() in ('*', '/'):
            op = consume()
            right = factor()
            left = (op, left, right)
        return left

    def factor():
        t = peek()
        if t is None:
            raise FormulaError("公式意外结束")
        if t == '(':
            consume()
            result = expr()
            if peek() != ')':
                raise FormulaError("缺少右括号 ')'")
            consume()
            return result
        if t == '-':
            consume()
            return ('-', ('num', 0.0), factor())
        if t in ('+', '*', '/'):
            raise FormulaError(f"意外的运算符 '{t}'")
        # Number literal
        try:
            val = float(t)
            consume()
            return ('num', val)
        except ValueError:
            pass
        # Column name reference
        consume()
        return ('ref', t)

    ast = expr()
    if peek() is not None:
        raise FormulaError(f"公式末尾存在意外字符 '{peek()}'")
    return ast


def collect_refs(ast):
    """Collect all column name references from an AST."""
    refs = []
    if ast[0] == 'ref':
        refs.append(ast[1])
    elif ast[0] == 'num':
        pass
    else:
        # Binary or unary operator
        if len(ast) == 3:
            refs.extend(collect_refs(ast[1]))
            refs.extend(collect_refs(ast[2]))
        elif len(ast) == 2:
            refs.extend(collect_refs(ast[1]))
    return refs


def evaluate_formula(ast, values):
    """Evaluate a formula AST against a values dict (name -> float).

    None values are treated as 0.0.  Division by zero raises FormulaError.
    Result is rounded to 2 decimal places.
    """
    if ast[0] == 'num':
        return ast[1]
    if ast[0] == 'ref':
        v = values.get(ast[1], 0.0)
        if v is None:
            v = 0.0
        return v
    # Unary minus: ('-', ('num', 0.0), operand)
    if ast[0] == '-' and len(ast) == 3 and ast[1] == ('num', 0.0):
        return -evaluate_formula(ast[2], values)
    # Binary operators
    op = ast[0]
    left = evaluate_formula(ast[1], values)
    right = evaluate_formula(ast[2], values)
    if op == '+':
        return left + right
    if op == '-':
        return left - right
    if op == '*':
        return left * right
    if op == '/':
        if right == 0:
            raise FormulaError("除零错误")
        return left / right
    raise FormulaError(f"未知运算符 '{op}'")


def split_row(source_row, ref_by_employee, source_headers, reference_headers, config,
              source_col_map, ref_col_map, splitting_col_set):
    """Split a row based on reference data."""
    # Precompute all column indices once using O(1) map lookups
    employee_id_col = source_col_map[config['input']['sheet']['source']['columns']['employee_id']]
    ref_hours_col = ref_col_map[config['input']['sheet']['reference']['columns']['project_hours']]
    project_id_col = source_col_map.get(config['input']['sheet']['source']['columns']['project_id'])
    project_category_col = source_col_map.get(config['input']['sheet']['source']['columns']['project_category'])
    project_hours_col = source_col_map.get(config['input']['sheet']['source']['columns']['project_hours'])
    ref_project_id_col = ref_col_map.get(config['input']['sheet']['reference']['columns']['project_id'])
    ref_project_category_col = ref_col_map.get(config['input']['sheet']['reference']['columns']['project_category'])
    ref_project_hours_col = ref_col_map.get(config['input']['sheet']['reference']['columns']['project_hours'])

    employee_id = source_row[employee_id_col - 1].value

    # O(1) lookup in pre-indexed reference dict
    matching_ref_rows = ref_by_employee.get(employee_id, [])

    if not matching_ref_rows:
        # No matching reference rows, copy source row as is
        return [list(source_row)]

    # 计算reference表中匹配行的总工时
    total_ref_hours = 0
    for ref_row in matching_ref_rows:
        try:
            ref_hours = float(ref_row[ref_hours_col - 1].value)
        except (ValueError, TypeError):
            fatal(f"Error: Non-numeric project_hours '{ref_row[ref_hours_col - 1].value}' "
                  f"in reference sheet for employee_id='{employee_id}'")
        total_ref_hours += ref_hours

    if total_ref_hours == 0:
        # 工时为0时，不拆分金额，但用 reference 的 project_id/project_category/project_hours 替换源表对应列
        ref_row = matching_ref_rows[0]
        result_row = list(source_row)
        if project_id_col and ref_project_id_col:
            result_row[project_id_col - 1].value = ref_row[ref_project_id_col - 1].value
        if project_category_col and ref_project_category_col:
            result_row[project_category_col - 1].value = ref_row[ref_project_category_col - 1].value
        if project_hours_col and ref_project_hours_col:
            result_row[project_hours_col - 1].value = ref_row[ref_project_hours_col - 1].value
        return [result_row]

    # Split the row
    # Track remaining values for each splitting column (avoid deepcopy of Cell objects)
    remain_values = []
    for j, cell in enumerate(source_row):
        if cell.value is not None and cell.column in splitting_col_set:
            try:
                remain_values.append(float(cell.value))
            except (ValueError, TypeError):
                fatal(f"Error: 无法拆分'{source_headers[j]}:{source_row[j].value}'")
        else:
            remain_values.append(0.0)  # placeholder for non-splitting columns

    result_rows = []
    for i, ref_row in enumerate(matching_ref_rows):
        try:
            ref_hours = float(ref_row[ref_hours_col - 1].value)
        except (ValueError, TypeError):
            fatal(f"Error: Non-numeric project_hours '{ref_row[ref_hours_col - 1].value}' "
                  f"in reference sheet for employee_id='{employee_id}'")
        ratio = ref_hours / total_ref_hours  # 根据reference表中的总工时计算比例

        # Create new row with same style as source
        new_row = []
        for j, cell in enumerate(source_row):
            new_cell = copy(cell)
            if cell.value is not None and cell.column in splitting_col_set:
                # Split numeric values
                try:
                    if i < len(matching_ref_rows) - 1:
                        proportional = float(cell.value) * ratio
                        # 向零取整：正数向下、负数向上，确保余数不反号
                        if proportional >= 0:
                            new_cell.value = math.floor(proportional * 100) / 100
                        else:
                            new_cell.value = math.ceil(proportional * 100) / 100
                        remain_values[j] -= new_cell.value
                    else:
                        new_cell.value = round(remain_values[j], 2)
                except (ValueError, TypeError):
                    fatal(f"Error: 无法拆分'{source_headers[j]}:{source_row[j].value}'")
            new_row.append(new_cell)

        # 只更新project_id, project_category, project_hours这三列
        if project_id_col and ref_project_id_col:
            new_row[project_id_col - 1].value = ref_row[ref_project_id_col - 1].value


        if project_category_col and ref_project_category_col:
            new_row[project_category_col - 1].value = ref_row[ref_project_category_col - 1].value

        if project_hours_col and ref_project_hours_col:
            new_row[project_hours_col - 1].value = ref_row[ref_project_hours_col - 1].value

        result_rows.append(new_row)

    return result_rows

def populate_payment_account(rows, payment_mapping, config, source_col_map, source_headers):
    """Populate payment_account column for each row (without merging)."""
    employer_name_col = source_col_map.get(config['input']['sheet']['source']['columns']['employer_name'])
    project_id_col = source_col_map.get(config['input']['sheet']['source']['columns']['project_id'])
    payment_account_col = source_col_map.get(config['input']['sheet']['source']['columns']['payment_account'])
    employee_id_col = source_col_map.get(config['input']['sheet']['source']['columns']['employee_id'])

    # Ensure all rows have enough cells for payment_account column
    # (source sheet may not have the payment_account column)
    max_cols = len(source_headers)
    if payment_account_col:
        for row in rows:
            while len(row) < max_cols:
                dummy = Cell(None, column=len(row) + 1)
                dummy.value = None
                dummy._style = copy(row[0]._style) if row else None
                row.append(dummy)

    for row in rows:
        proj_id = row[project_id_col - 1].value
        proj_id_str = str(proj_id).strip() if proj_id is not None else ''
        employer_name = row[employer_name_col - 1].value if employer_name_col else None
        employer_name_str = str(employer_name).strip() if employer_name is not None else ''
        emp_id = row[employee_id_col - 1].value

        # Use composite key (employer_name, project_id) to look up payment_account
        lookup_key = (employer_name_str, proj_id_str)
        if lookup_key not in payment_mapping:
            fatal(f"Error: employee_id='{emp_id}', employer_name='{employer_name_str}', "
                  f"project_id='{proj_id_str}' has no matching payment account")

        if payment_account_col:
            row[payment_account_col - 1].value = payment_mapping[lookup_key]


def merge_rows_by_account(split_rows, source_headers, config,
                          source_col_map, splitting_col_list):
    """Merge split result rows by (employee_id, payment_account)."""
    if not split_rows:
        return []

    # Precompute all column indices once using O(1) map lookups
    employee_id_col = source_col_map.get(config['input']['sheet']['source']['columns']['employee_id'])
    project_id_col = source_col_map.get(config['input']['sheet']['source']['columns']['project_id'])
    project_category_col = source_col_map.get(config['input']['sheet']['source']['columns']['project_category'])
    project_hours_col = source_col_map.get(config['input']['sheet']['source']['columns']['project_hours'])
    payment_account_col = source_col_map.get(config['input']['sheet']['source']['columns']['payment_account'])

    splitting_col_indices = splitting_col_list  # Already precomputed

    # Group rows by (employee_id, payment_account), preserving first-occurrence order
    # Dict maintains insertion order (Python 3.7+)
    groups = {}  # group_key -> [rows]

    for row in split_rows:
        emp_id = row[employee_id_col - 1].value
        proj_account = row[payment_account_col - 1].value if payment_account_col else None
        group_key = (emp_id, proj_account)

        if group_key in groups:
            groups[group_key].append(row)
        else:
            groups[group_key] = [row]

    # Merge each group
    merged_rows = []
    for group_key, rows in groups.items():
        emp_id, proj_account = group_key

        # Start with shallow copy of first row (list of cells)
        merged_row = list(rows[0])

        # Collect distinct project_ids and project_categories
        proj_ids = []
        proj_categories = []
        total_hours = 0.0

        # Initialize splitting column sums
        splitting_sums = {col_idx: 0.0 for col_idx in splitting_col_indices if col_idx}

        for row in rows:
            pid = row[project_id_col - 1].value
            pid_str = str(pid).strip() if pid is not None else ''
            if pid_str and pid_str not in proj_ids:
                proj_ids.append(pid_str)

            if project_category_col:
                cat = row[project_category_col - 1].value
                cat_str = str(cat).strip() if cat is not None else ''
                if cat_str and cat_str not in proj_categories:
                    proj_categories.append(cat_str)

            if project_hours_col:
                hours = row[project_hours_col - 1].value
                try:
                    total_hours += float(hours) if hours is not None else 0.0
                except (ValueError, TypeError):
                    fatal(f"Error: Cannot sum project_hours value '{hours}'")

            # Sum splitting columns
            for col_idx in splitting_col_indices:
                if col_idx:
                    val = row[col_idx - 1].value
                    try:
                        splitting_sums[col_idx] += float(val) if val is not None else 0.0
                    except (ValueError, TypeError):
                        fatal(f"Error: Cannot sum splitting column '{source_headers[col_idx - 1]}' value '{val}'")

        # Set merged values
        merged_row[project_id_col - 1].value = ','.join(proj_ids)
        if project_category_col:
            merged_row[project_category_col - 1].value = ','.join(proj_categories)
        if project_hours_col:
            merged_row[project_hours_col - 1].value = round(total_hours, 2)
        if payment_account_col:
            merged_row[payment_account_col - 1].value = proj_account

        # Set summed splitting columns
        for col_idx, sum_val in splitting_sums.items():
            merged_row[col_idx - 1].value = round(sum_val, 2)

        merged_rows.append(merged_row)

    return merged_rows


def validate_sheets(config, wb):
    """Validate that required sheets and columns exist, and run pre-processing data checks."""
    source_sheet = config['input']['sheet']['source']['name']
    reference_sheet = config['input']['sheet']['reference']['name']

    # Check if sheets exist
    if source_sheet not in wb.sheetnames:
        fatal(f"Error: Source sheet '{source_sheet}' does not exist")
    if reference_sheet not in wb.sheetnames:
        fatal(f"Error: Reference sheet '{reference_sheet}' does not exist")

    source = wb[source_sheet]
    reference = wb[reference_sheet]

    # Get header rows
    source_headers = [cell.value for cell in source[1]]
    reference_headers = [cell.value for cell in reference[1]]

    # Collect all errors first, report them together at the end
    errors = []

    # Payment sheet is required
    payment_sheet = config['input']['sheet']['payment']['name']
    if payment_sheet not in wb.sheetnames:
        fatal(f"Error: Payment sheet '{payment_sheet}' does not exist")
    payment = wb[payment_sheet]
    payment_headers = [cell.value for cell in payment[1]]

    # Validate required columns in source sheet
    required_source_columns = {
        'employee_id': config['input']['sheet']['source']['columns']['employee_id']
    }
    for col_id, col_name in required_source_columns.items():
        if col_name not in source_headers:
            fatal(f"Error: Required column '{col_name}' not found in source sheet")

    # Validate splitting columns in source sheet
    for col_name in config['input']['splitting_columns']:
        if col_name not in source_headers:
            fatal(f"Error: Splitting column '{col_name}' not found in source sheet")

    # Validate computed columns (if configured)
    computed_columns_cfg = config.get('input', {}).get('computed_columns', {})
    if computed_columns_cfg:
        splitting_col_set = set(config['input']['splitting_columns'])
        computed_names = list(computed_columns_cfg.keys())
        known = set(source_headers) | set(computed_names)
        for i, name in enumerate(computed_names):
            formula = computed_columns_cfg[name]
            # Check computed column name exists in source headers
            if name not in source_headers:
                errors.append(f"计算列 '{name}' 不存在于源表中")
            # Parse formula and validate references
            try:
                tokens = tokenize_formula(formula, known)
                ast = parse_formula(tokens)
            except FormulaError as e:
                errors.append(f"计算列 '{name}' 的公式无效: {e}")
                continue
            # Check all referenced columns: must be in splitting_columns or an earlier computed column
            for ref in collect_refs(ast):
                if ref in splitting_col_set:
                    continue  # OK: reference is a splitting column
                if ref in computed_names[:i]:
                    continue  # OK: reference is an earlier computed column
                if ref in source_headers:
                    errors.append(
                        f"计算列 '{name}' 公式中引用的 '{ref}' 不在 splitting_columns 中"
                    )
                else:
                    errors.append(
                        f"计算列 '{name}' 公式中引用的 '{ref}' 不存在于源表列中"
                    )

    # Validate required columns in reference sheet
    required_ref_columns = {
        'employee_id': config['input']['sheet']['reference']['columns']['employee_id'],
        'project_id': config['input']['sheet']['reference']['columns']['project_id'],
        'project_category': config['input']['sheet']['reference']['columns']['project_category'],
        'project_hours': config['input']['sheet']['reference']['columns']['project_hours']
    }
    for col_id, col_name in required_ref_columns.items():
        if col_name not in reference_headers:
            fatal(f"Error: Required column '{col_name}' not found in reference sheet")

    # Validate required columns in payment sheet
    required_payment_columns = {
        'project_id': config['input']['sheet']['payment']['columns']['project_id'],
        'payment_account': config['input']['sheet']['payment']['columns']['payment_account']
    }
    for col_id, col_name in required_payment_columns.items():
        if col_name not in payment_headers:
            fatal(f"Error: Required column '{col_name}' not found in payment sheet")

    # --- Pre-processing data validation ---
    ref_employee_id_col = get_column_index(reference_headers,
                                           config['input']['sheet']['reference']['columns']['employee_id'])
    ref_project_id_col = get_column_index(reference_headers,
                                          config['input']['sheet']['reference']['columns']['project_id'])
    ref_hours_col = get_column_index(reference_headers,
                                     config['input']['sheet']['reference']['columns']['project_hours'])

    source_project_id_col = get_column_index(source_headers,
                                              config['input']['sheet']['source']['columns']['project_id'])

    # Payment-related column indices
    payment_mapping = {}
    payment_project_id_col = get_column_index(payment_headers,
                                              config['input']['sheet']['payment']['columns']['project_id'])
    payment_payment_account_col = get_column_index(payment_headers,
                                                    config['input']['sheet']['payment']['columns']['payment_account'])
    payment_employer_name_col = get_column_index(payment_headers,
                                                  config['input']['sheet']['payment']['columns']['employer_name'])

    # --- Payment checks ---
    # Build payment_mapping and payment_project_ids before reference/source scans
    payment_project_ids = set()
    ref_employee_ids = set()
    ref_projects_by_employee = {}  # employee_id -> set of project_ids (for composite key check)
    ref_hours_by_employee = {}  # employee_id -> total_hours (to detect 0-hour rows)

    # Check 2 & 3: Payment table - no duplicate (employer_name, project_id) + payment_account not empty
    seen_payment_pairs = set()
    reported_duplicates = set()
    for row in payment.iter_rows(min_row=2):
        proj_id = row[payment_project_id_col - 1].value
        proj_account = row[payment_payment_account_col - 1].value
        employer_name = row[payment_employer_name_col - 1].value

        # Skip rows with empty project_id or empty employer_name
        if proj_id is None or str(proj_id).strip() == '':
            continue
        if employer_name is None or str(employer_name).strip() == '':
            continue

        proj_id_str = str(proj_id).strip()
        employer_name_str = str(employer_name).strip()

        # Check 2: no duplicate (employer_name, project_id) (report each duplicate pair only once)
        pair = (employer_name_str, proj_id_str)
        if pair in seen_payment_pairs:
            if pair not in reported_duplicates:
                errors.append(f"Duplicate (employer_name='{employer_name_str}', project_id='{proj_id_str}') "
                             f"found in sheet '{payment_sheet}'")
                reported_duplicates.add(pair)
            continue
        seen_payment_pairs.add(pair)

        # Check 3: payment_account must not be empty
        if proj_account is None or str(proj_account).strip() == '':
            errors.append(f"project_id '{proj_id_str}' has empty payment_account in "
                         f"payment sheet '{payment_sheet}'")
            continue

        payment_mapping[pair] = str(proj_account).strip()

    # Build set of all project_ids in payment for Check 4 partial check
    payment_project_ids = set(pid for (_, pid) in payment_mapping.keys())

    # --- Consolidated single-pass scan of reference sheet ---
    # Checks 1, 4, 6, 8, 9, and partial 10 (type collection)
    ref_pairs = set()
    reported_ref_duplicates = set()
    ref_missing_pids = set()
    reported_empty_ref_id = False
    ref_id_types = set()

    for row_num, row in enumerate(reference.iter_rows(min_row=2), start=2):
        emp_id = row[ref_employee_id_col - 1].value
        proj_id = row[ref_project_id_col - 1].value
        hours_val = row[ref_hours_col - 1].value

        # Check 6: empty employee_id
        if not reported_empty_ref_id and (emp_id is None or
           (isinstance(emp_id, str) and str(emp_id).strip() == '')):
            errors.append(f"Empty employee_id in reference sheet '{reference_sheet}' row {row_num}")
            reported_empty_ref_id = True

        # Check 10: type info (skip if empty)
        if emp_id is not None and not (isinstance(emp_id, str) and str(emp_id).strip() == ''):
            ref_id_types.add(type(emp_id).__name__)

        # Build ref_employee_ids and ref_projects_by_employee for source scan checks
        if emp_id is not None:
            ref_employee_ids.add(emp_id)
            if emp_id not in ref_projects_by_employee:
                ref_projects_by_employee[emp_id] = set()
            if proj_id is not None and str(proj_id).strip() != '':
                ref_projects_by_employee[emp_id].add(str(proj_id).strip())

        # Track total hours per employee (for Check 11 zero-hour detection)
        if emp_id is not None:
            if emp_id not in ref_hours_by_employee:
                ref_hours_by_employee[emp_id] = 0.0
            if hours_val is not None:
                try:
                    ref_hours_by_employee[emp_id] += float(hours_val)
                except (ValueError, TypeError):
                    pass  # Non-numeric hours caught by Check 8

        # Check 1: duplicate (employee_id, project_id)
        if emp_id is not None and proj_id is not None:
            pair = (emp_id, proj_id)
            if pair in ref_pairs:
                if pair not in reported_ref_duplicates:
                    errors.append(f"Duplicate (employee_id='{emp_id}', project_id='{proj_id}') "
                                  f"found in reference sheet '{reference_sheet}'")
                    reported_ref_duplicates.add(pair)
            else:
                ref_pairs.add(pair)

        # Check 4: project_id in payment
        if proj_id is not None and str(proj_id).strip() != '':
            proj_id_str = str(proj_id).strip()
            if proj_id_str not in payment_project_ids:
                ref_missing_pids.add(proj_id_str)

        # Checks 8 and 9: non-numeric and negative hours
        if hours_val is not None:
            try:
                hours_float = float(hours_val)
                if hours_float < 0:
                    errors.append(f"Negative project_hours '{hours_val}' in reference sheet "
                                  f"'{reference_sheet}' row {row_num} "
                                  f"(employee_id='{emp_id}', project_id='{proj_id}')")
            except (ValueError, TypeError):
                errors.append(f"Non-numeric project_hours '{hours_val}' in reference sheet "
                              f"'{reference_sheet}' row {row_num} "
                              f"(employee_id='{emp_id}', project_id='{proj_id}')")

    # Report Check 4 errors after reference scan
    if ref_missing_pids:
        for pid in sorted(ref_missing_pids):
            errors.append(f"project_id '{pid}' in reference sheet '{reference_sheet}' "
                         f"has no matching record in sheet '{payment_sheet}'")

    # Check 0h: Zero-hour employees must have at most one distinct project_id in reference
    for emp_id, total_hours in ref_hours_by_employee.items():
        if total_hours == 0.0:
            projects = ref_projects_by_employee.get(emp_id, set())
            if len(projects) > 1:
                errors.append(f"employee_id='{emp_id}' has total hours=0 but multiple project_ids "
                             f"in reference sheet '{reference_sheet}': {sorted(projects)}. "
                             f"When hours are 0, only one project_id is allowed.")

    # --- Consolidated single-pass scan of source sheet ---
    # Checks 5, 7, partial 10 (type collection), and collect employer names for split rows
    src_employee_id_col = get_column_index(source_headers,
                                           config['input']['sheet']['source']['columns']['employee_id'])
    reported_empty_src_id = False
    src_id_types = set()
    source_missing_pairs = set()
    src_employers_by_employee = {}  # employee_id -> set(employer_names) for split-eligible rows

    src_employer_name_col = get_column_index(source_headers,
                                              config['input']['sheet']['source']['columns']['employer_name'])

    for row_num, row in enumerate(source.iter_rows(min_row=2), start=2):
        emp_id = row[src_employee_id_col - 1].value

        # Check 7: empty employee_id
        if not reported_empty_src_id and (emp_id is None or
           (isinstance(emp_id, str) and str(emp_id).strip() == '')):
            errors.append(f"Empty employee_id in source sheet '{source_sheet}' row {row_num}")
            reported_empty_src_id = True

        # Check 10: type info
        if emp_id is not None and not (isinstance(emp_id, str) and str(emp_id).strip() == ''):
            src_id_types.add(type(emp_id).__name__)

        # Check 5: non-split rows must have payment mapping
        if source_project_id_col and src_employer_name_col:
            if emp_id in ref_employee_ids:
                # Collect all employer_names for composite key check (Check 11)
                employer_name = row[src_employer_name_col - 1].value
                employer_name_str = str(employer_name).strip() if employer_name is not None else ''
                if emp_id not in src_employers_by_employee:
                    src_employers_by_employee[emp_id] = set()
                src_employers_by_employee[emp_id].add(employer_name_str)
                continue  # Will be split/replaced; source project_id replaced by reference
            proj_id = row[source_project_id_col - 1].value
            employer_name = row[src_employer_name_col - 1].value
            if proj_id is not None and str(proj_id).strip() != '' \
               and employer_name is not None and str(employer_name).strip() != '':
                pair = (str(employer_name).strip(), str(proj_id).strip())
                if pair not in payment_mapping:
                    source_missing_pairs.add(pair)

    # Report Check 5 errors after source scan
    if source_missing_pairs:
        for pair in sorted(source_missing_pairs):
            errors.append(f"(employer_name='{pair[0]}', project_id='{pair[1]}') in source sheet "
                         f"'{source_sheet}' has no matching record in sheet '{payment_sheet}'")

    # Check 11: Split-eligible rows - composite key (employer_name, project_id) must exist in payment
    # Reference project_ids are used after split (and also for zero-hour rows which now
    # replace source project_id with reference project_id).
    split_missing_pairs = set()
    for emp_id, employer_names in src_employers_by_employee.items():
        ref_projects = ref_projects_by_employee.get(emp_id, set())
        for employer_name in employer_names:
            for proj_id in ref_projects:
                pair = (employer_name, proj_id)
                if pair not in payment_mapping:
                    split_missing_pairs.add((emp_id, employer_name, proj_id))
    for emp_id, employer_name, proj_id in sorted(split_missing_pairs):
        errors.append(f"employee_id='{emp_id}', employer_name='{employer_name}', "
                     f"project_id='{proj_id}' has no matching record "
                     f"in sheet '{payment_sheet}'")

    # Check 10: employee_id type mismatch between sheets
    all_id_types = ref_id_types | src_id_types
    if len(all_id_types) > 1:
        types_str = ', '.join(sorted(all_id_types))
        errors.append(f"employee_id type mismatch: found types [{types_str}] across "
                     f"source and reference sheets. All employee_id values should be the same type.")

    # Report all collected errors at once
    if errors:
        error_msg = "Validation errors found:\n" + "\n".join(f"  - {e}" for e in errors)
        fatal(error_msg)

    return source, reference, source_headers, reference_headers, payment_mapping


def verify_output(config, source_headers):
    """Verify the output file after processing. Returns a list of error messages."""
    errors = []
    output_path = config['output']['path']
    result_sheet_name = config['output']['sheet']['result']['name']
    source_sheet_name = config['input']['sheet']['source']['name']
    splitting_columns = config['input']['splitting_columns']

    splitting_columns = list(dict.fromkeys(config['input']['splitting_columns']))
    computed_columns_cfg = config.get('input', {}).get('computed_columns', {})
    computed_col_names = list(computed_columns_cfg.keys()) if computed_columns_cfg else []

    if not os.path.exists(output_path):
        errors.append(f"Output file '{output_path}' does not exist")
        return errors

    try:
        wb = load_workbook(output_path)
    except Exception as e:
        errors.append(f"Cannot open output file '{output_path}': {e}")
        return errors

    # Check result sheet exists
    if result_sheet_name not in wb.sheetnames:
        errors.append(f"Result sheet '{result_sheet_name}' not found in output file")
        wb.close()
        return errors

    result = wb[result_sheet_name]

    # Check source sheet exists in output (output is a copy of input)
    if source_sheet_name not in wb.sheetnames:
        errors.append(f"Source sheet '{source_sheet_name}' not found in output file")
        wb.close()
        return errors

    source = wb[source_sheet_name]
    result_headers = [cell.value for cell in result[1]]
    source_out_headers = [cell.value for cell in source[1]]

    # Check no empty rows in result
    for i, row in enumerate(result.iter_rows(min_row=2), start=2):
        if all(cell.value is None for cell in row):
            errors.append(f"Empty row {i} in result sheet '{result_sheet_name}'")

    # Check splitting columns: numeric in result
    # Precompute column indices for all splitting columns
    result_col_map_verify = {}
    for col_name in splitting_columns:
        col_idx = get_column_index(result_headers, col_name)
        if col_idx is None:
            errors.append(f"Splitting column '{col_name}' not found in result headers")
        else:
            result_col_map_verify[col_name] = col_idx

    # Also check computed columns are numeric
    for col_name in computed_col_names:
        col_idx = get_column_index(result_headers, col_name)
        if col_idx is None:
            errors.append(f"Computed column '{col_name}' not found in result headers")
        elif col_name not in result_col_map_verify:
            result_col_map_verify[col_name] = col_idx

    # Single pass through result: validate numeric for all splitting + computed columns
    for i, row in enumerate(result.iter_rows(min_row=2), start=2):
        for col_name, col_idx in result_col_map_verify.items():
            val = row[col_idx - 1].value
            if val is not None:
                try:
                    float(val)
                except (ValueError, TypeError):
                    errors.append(f"Non-numeric value '{val}' in '{col_name}' at result row {i}")

    # Grand total consistency: sum of each splitting column in result should
    # equal the sum in source (within rounding tolerance)
    # Precompute column indices
    src_col_map_verify = {}
    res_col_map_verify = {}
    for col_name in splitting_columns:
        src_idx = get_column_index(source_out_headers, col_name)
        res_idx = get_column_index(result_headers, col_name)
        if src_idx and res_idx:
            src_col_map_verify[col_name] = src_idx
            res_col_map_verify[col_name] = res_idx

    if src_col_map_verify:
        # Single pass through source: sum all splitting columns
        src_sums = {col_name: 0.0 for col_name in src_col_map_verify}
        for row in source.iter_rows(min_row=2):
            for col_name, col_idx in src_col_map_verify.items():
                val = row[col_idx - 1].value
                if val is not None:
                    try:
                        src_sums[col_name] += float(val)
                    except (ValueError, TypeError):
                        pass  # Non-numeric source values were caught by pre-checks

        # Single pass through result: sum all splitting columns
        res_sums = {col_name: 0.0 for col_name in res_col_map_verify}
        for row in result.iter_rows(min_row=2):
            for col_name, col_idx in res_col_map_verify.items():
                val = row[col_idx - 1].value
                if val is not None:
                    try:
                        res_sums[col_name] += float(val)
                    except (ValueError, TypeError):
                        pass  # Already caught above

        # Compare each column's sums
        for col_name in splitting_columns:
            if col_name not in src_col_map_verify or col_name not in res_col_map_verify:
                continue
            src_sum = src_sums.get(col_name, 0.0)
            res_sum = res_sums.get(col_name, 0.0)
            if abs(res_sum - src_sum) > 0.001:
                errors.append(f"Total mismatch for '{col_name}': "
                             f"source sum={src_sum:.2f}, result sum={res_sum:.2f}")

    wb.close()
    return errors


def validate_config(config):
    """Validate configuration structure. Returns (merge_by_payment_account, computed_columns)."""
    errors = []

    if config is None:
        fatal("Error: Configuration is empty")

    # Check input section
    inp = config.get('input')
    if inp is None:
        errors.append("Missing 'input' section in configuration")
    else:
        if 'path' not in inp or not isinstance(inp['path'], str) or inp['path'].strip() == '':
            errors.append("Missing or invalid 'input.path' in configuration")

        sheets = inp.get('sheet')
        if sheets is None:
            errors.append("Missing 'input.sheet' section in configuration")
        else:
            # Source sheet
            source = sheets.get('source')
            if source is None:
                errors.append("Missing 'input.sheet.source' in configuration")
            else:
                if 'name' not in source or not isinstance(source['name'], str) or source['name'].strip() == '':
                    errors.append("Missing or invalid 'input.sheet.source.name' in configuration")
                if 'columns' not in source:
                    errors.append("Missing 'input.sheet.source.columns' in configuration")
                elif 'employee_id' not in source['columns']:
                    errors.append("Missing 'input.sheet.source.columns.employee_id' in configuration")

            # Reference sheet
            ref = sheets.get('reference')
            if ref is None:
                errors.append("Missing 'input.sheet.reference' in configuration")
            else:
                if 'name' not in ref or not isinstance(ref['name'], str) or ref['name'].strip() == '':
                    errors.append("Missing or invalid 'input.sheet.reference.name' in configuration")
                if 'columns' not in ref:
                    errors.append("Missing 'input.sheet.reference.columns' in configuration")
                else:
                    for col in ['employee_id', 'project_id', 'project_category', 'project_hours']:
                        if col not in ref['columns']:
                            errors.append(f"Missing 'input.sheet.reference.columns.{col}' in configuration")

            # Payment sheet (required)
            pay = sheets.get('payment')
            if pay is None:
                errors.append("Missing 'input.sheet.payment' section in configuration")
            elif not ('name' in pay and isinstance(pay['name'], str) and pay['name'].strip() != '' \
               and 'columns' in pay \
               and 'project_id' in pay['columns'] \
               and 'payment_account' in pay['columns'] \
               and 'employer_name' in pay['columns']):
                errors.append("Missing or incomplete 'input.sheet.payment' configuration")

            # employer_name and payment_account are required in source columns
            source_cols = sheets.get('source', {}).get('columns', {})
            if 'employer_name' not in source_cols:
                errors.append("'employer_name' is required in source columns")
            if 'payment_account' not in source_cols:
                errors.append("'payment_account' is required in source columns")

            # Read merge_by_payment_account config (default: False)
            merge_by_payment_account = config.get('merge_by_payment_account', False)

        # splitting_columns
        split_cols = inp.get('splitting_columns')
        if split_cols is None:
            errors.append("Missing 'input.splitting_columns' in configuration")
        elif not isinstance(split_cols, list) or len(split_cols) == 0:
            errors.append("'input.splitting_columns' must be a non-empty list")
        else:
            seen = set()
            for col in split_cols:
                if col in seen:
                    errors.append(f"Duplicate entry '{col}' in 'input.splitting_columns'")
                else:
                    seen.add(col)

        # computed_columns (optional)
        computed_columns = inp.get('computed_columns')
        if computed_columns is not None:
            if not isinstance(computed_columns, dict) or len(computed_columns) == 0:
                errors.append("'input.computed_columns' must be a non-empty mapping")
            else:
                split_col_set = set(split_cols) if split_cols else set()
                for name, formula in computed_columns.items():
                    if not isinstance(name, str) or name.strip() == '':
                        errors.append("computed_columns key must be a non-empty string")
                    if not isinstance(formula, str) or formula.strip() == '':
                        errors.append(f"computed_columns '{name}' formula must be a non-empty string")
                    if name in split_col_set:
                        errors.append(f"计算列 '{name}' 不能同时出现在 input.splitting_columns 中")
                # Check for duplicate computed column names handled by dict (last wins),
                # but we warn if the same name is defined multiple times — done via dict,
                # so duplicates are naturally removed. No action needed.
        else:
            computed_columns = {}

    # Check output section
    out = config.get('output')
    if out is None:
        errors.append("Missing 'output' section in configuration")
    else:
        if 'path' not in out or not isinstance(out['path'], str) or out['path'].strip() == '':
            errors.append("Missing or invalid 'output.path' in configuration")
        out_sheet = out.get('sheet')
        if out_sheet is None:
            errors.append("Missing 'output.sheet' in configuration")
        elif 'result' not in out_sheet:
            errors.append("Missing 'output.sheet.result' in configuration")
        elif 'name' not in out_sheet['result'] or not isinstance(out_sheet['result']['name'], str) \
                or out_sheet['result']['name'].strip() == '':
            errors.append("Missing or invalid 'output.sheet.result.name' in configuration")

    if errors:
        error_msg = "Configuration errors:\n" + "\n".join(f"  - {e}" for e in errors)
        fatal(error_msg)

    return merge_by_payment_account, computed_columns


def process_excel(config):
    """Process Excel file according to configuration."""
    # Validate config structure first
    merge_by_payment_account, computed_columns_cfg = validate_config(config)

    input_path = config['input']['path']
    output_path = config['output']['path']
    result_sheet = config['output']['sheet']['result']['name']

    # Check if input file exists
    if not os.path.exists(input_path):
        fatal(f"Error: Input file '{input_path}' does not exist")

    try:
        # Load the workbook
        print(f"正在加载工作簿: {input_path}")
        wb = load_workbook(input_path, data_only=True)
        print("工作簿加载完成")
        
        # Validate sheets and columns
        source, reference, source_headers, reference_headers, payment_mapping = \
            validate_sheets(config, wb)
        print("验证完成")

        # Create a new workbook for output
        output_wb = load_workbook(input_path)

        # If result sheet exists, remove it
        if result_sheet in output_wb.sheetnames:
            output_wb.remove(output_wb[result_sheet])

        keep_style = config.get('keep_style', True)

        # Create new sheet
        result = output_wb.create_sheet(result_sheet)

        # Copy headers
        for cell in source[1]:
            new_cell = result.cell(row=1, column=cell.column)
            new_cell.value = cell.value
            if keep_style:
                copy_cell_style(cell, new_cell)

        # Ensure payment_account column exists in source_headers (payment is required)
        payment_account_name = config['input']['sheet']['source']['columns'].get('payment_account')
        if payment_account_name and get_column_index(source_headers, payment_account_name) is None:
            new_col = len(source_headers) + 1
            header_cell = result.cell(row=1, column=new_col)
            header_cell.value = payment_account_name
            source_headers.append(payment_account_name)

        # Precompute column name -> index maps for O(1) lookups (after source_headers is finalized)
        source_col_map = {name: idx + 1 for idx, name in enumerate(source_headers)}
        ref_col_map = {name: idx + 1 for idx, name in enumerate(reference_headers)}

        # Precompute splitting column indices as set (fast membership test in split_row)
        # and ordered list (for merge_rows_by_account)
        splitting_col_set = set()
        splitting_col_list = []
        for col_name in config['input']['splitting_columns']:
            idx = source_col_map.get(col_name)
            if idx is not None:
                splitting_col_set.add(idx)
                if idx not in splitting_col_list:
                    splitting_col_list.append(idx)

        # Pre-parse computed column formulas
        computed_columns_cfg = config.get('input', {}).get('computed_columns', {})
        computed_asts = []  # list of (name, ast, col_index)
        if computed_columns_cfg:
            known_for_parse = set(source_headers)
            for cc_name, cc_formula in computed_columns_cfg.items():
                ast = parse_formula(tokenize_formula(cc_formula, known_for_parse))
                cc_idx = source_col_map[cc_name]
                computed_asts.append((cc_name, ast, cc_idx))
                known_for_parse.add(cc_name)

        # Process data rows
        current_row = 2
        reference_rows = list(reference.iter_rows(min_row=2))  # Convert iterator to list

        # Pre-index reference rows by employee_id for O(1) lookup
        ref_employee_id_col = ref_col_map[config['input']['sheet']['reference']['columns']['employee_id']]
        ref_by_employee = {}
        for ref_row in reference_rows:
            emp_id = ref_row[ref_employee_id_col - 1].value
            if emp_id not in ref_by_employee:
                ref_by_employee[emp_id] = []
            ref_by_employee[emp_id].append(ref_row)

        total_source_rows = source.max_row - 1  # Subtract header row
        total_reference_rows = len(reference_rows)
        print(f"开始处理数据，共 {total_source_rows} 行（参考表 {total_reference_rows} 行，{len(ref_by_employee)} 个员工）")
        write_batch_size = config.get('write_batch_size', 500)
        row_counter = 0
        all_result_rows = []  # Collect (values_list, cells_list) tuples, write later
        rows_to_compute = set()  # Indices of rows that need computed column evaluation
        src_row_groups = []  # List of (source_row, [result_indices]) for per-source-row verification

        # Pre-fetch column indices for split detection
        src_emp_col = source_col_map[config['input']['sheet']['source']['columns']['employee_id']]
        ref_hours_col = ref_col_map[config['input']['sheet']['reference']['columns']['project_hours']]

        t_total_start = time.time()
        t_split_total = 0.0
        t_merge_total = 0.0
        for row in source.iter_rows(min_row=2):
            row_counter += 1
            if row_counter % write_batch_size == 0 or row_counter == 1:
                elapsed = time.time() - t_total_start
                rate = row_counter / elapsed if elapsed > 0 else 0
                eta = (total_source_rows - row_counter) / rate if rate > 0 else 0
                print(f"正在处理第 {row_counter}/{total_source_rows} 行 (已耗时 {elapsed:.0f}s, 速度 {rate:.1f} 行/秒, 预计剩余 {eta:.0f}s)...")

            # Determine if this source row will be split (has matching ref with total hours > 0)
            emp_id_val = row[src_emp_col - 1].value
            matching = ref_by_employee.get(emp_id_val, [])
            total_ref = 0.0
            for r in matching:
                try:
                    total_ref += float(r[ref_hours_col - 1].value or 0)
                except (ValueError, TypeError):
                    pass
            is_split = len(matching) > 0 and total_ref > 0

            t0 = time.time()
            split_result_rows = split_row(row, ref_by_employee, source_headers, reference_headers, config,
                                           source_col_map, ref_col_map, splitting_col_set)
            t_split_total += time.time() - t0

            # Populate payment_account column for each split row
            populate_payment_account(split_result_rows, payment_mapping, config, source_col_map,
                                     source_headers)

            # Merge split rows by payment account (only if configured)
            if merge_by_payment_account:
                t0 = time.time()
                merged_rows = merge_rows_by_account(split_result_rows, source_headers, config,
                                                     source_col_map, splitting_col_list)
                t_merge_total += time.time() - t0
            else:
                merged_rows = split_result_rows

            start_idx = len(all_result_rows)
            for result_row in merged_rows:
                all_result_rows.append(result_row)

            # Mark split rows for computed column evaluation
            end_idx = len(all_result_rows)
            if is_split and computed_asts:
                for i in range(start_idx, end_idx):
                    rows_to_compute.add(i)
                src_row_groups.append((row, list(range(start_idx, end_idx))))

        total_elapsed = time.time() - t_total_start
        print(f"处理耗时分析：拆分 {t_split_total:.1f}s, 合并 {t_merge_total:.1f}s, 总计 {total_elapsed:.1f}s")

        # Compute formulas for split rows
        computed_errors = []
        if computed_asts and rows_to_compute:
            t_compute_start = time.time()
            max_cols = len(source_headers)
            # Pre-collect all column names referenced in any formula
            ref_names = set()
            for _cc_name, ast, _cc_idx in computed_asts:
                ref_names.update(collect_refs(ast))
            for idx in rows_to_compute:
                row = all_result_rows[idx]
                # Pad row to max_cols if needed (for newly appended columns)
                while len(row) < max_cols:
                    dummy = Cell(None, column=len(row) + 1)
                    dummy.value = None
                    row.append(dummy)
                # Build values dict from referenced columns only
                values = {}
                for hdr_name in ref_names:
                    col_idx = source_col_map[hdr_name]
                    v = row[col_idx - 1].value
                    if v is None:
                        values[hdr_name] = 0.0
                    else:
                        try:
                            values[hdr_name] = float(v)
                        except (ValueError, TypeError):
                            fatal(f"Error: 计算列引用的列 '{hdr_name}' 存在非数值 '{v}'")
                # Evaluate formulas in order
                for cc_name, ast, cc_idx in computed_asts:
                    result_val = evaluate_formula(ast, values)
                    result_val = round(result_val, 2)
                    values[cc_name] = result_val
                    row[cc_idx - 1].value = result_val
            t_compute_total = time.time() - t_compute_start
            print(f"计算列耗时：{t_compute_total:.1f}s (共 {len(rows_to_compute)} 行)")

            # Per-source-row verification: sum of computed columns across split rows
            # must equal the source row's original value.
            # Collect all mismatches first, then report after file is saved.
            src_employee_id_name = config['input']['sheet']['source']['columns']['employee_id']
            computed_errors = []
            for src_row, result_indices in src_row_groups:
                emp_id = src_row[source_col_map[src_employee_id_name] - 1].value
                for cc_name, _ast, cc_idx in computed_asts:
                    # Skip if source row doesn't have this column (e.g. newly appended)
                    if cc_idx > len(src_row):
                        continue
                    src_val = src_row[cc_idx - 1].value
                    if src_val is None:
                        continue  # Source had no value for this column, skip
                    try:
                        src_val_f = float(src_val)
                    except (ValueError, TypeError):
                        continue
                    # Sum computed values across all result rows from this source row
                    result_sum = 0.0
                    for ri in result_indices:
                        v = all_result_rows[ri][cc_idx - 1].value
                        if v is not None:
                            try:
                                result_sum += float(v)
                            except (ValueError, TypeError):
                                pass
                    if abs(result_sum - src_val_f) > 0.001:
                        computed_errors.append(
                            f"计算列 '{cc_name}' 源行 employee_id='{emp_id}' "
                            f"拆分结果合计 ({result_sum:.2f}) 与源行值 ({src_val_f:.2f}) 不一致"
                        )

        # Batch write to worksheet
        t_write_start = time.time()
        current_row = 2
        total_result_rows = len(all_result_rows)
        for batch_start in range(0, total_result_rows, write_batch_size):
            batch = all_result_rows[batch_start:batch_start + write_batch_size]
            for i, result_row in enumerate(batch):
                row_num = current_row + i
                for cell in result_row:
                    new_cell = result.cell(row=row_num, column=cell.column)
                    new_cell.value = cell.value
                    if keep_style:
                        copy_cell_style(cell, new_cell)
            current_row += len(batch)
        t_write_total = time.time() - t_write_start
        print(f"写入耗时：{t_write_total:.1f}s (共 {total_result_rows} 行，批次大小 {write_batch_size})")

        # Copy column dimensions
        for col in source.column_dimensions:
            result.column_dimensions[col].width = source.column_dimensions[col].width

        # Copy row dimensions
        for row in source.row_dimensions:
            result.row_dimensions[row].height = source.row_dimensions[row].height

        # Save the output file
        print("处理完成，正在保存输出文件...")
        output_wb.save(output_path)
        print("输出文件保存成功")

        # Report computed column verification errors (if any)
        if computed_errors:
            error_msg = "计算列校验错误（公式结果与源表不一致）:\n" + "\n".join(f"  - {e}" for e in computed_errors)
            fatal(error_msg)

        # Verify output after save
        print("正在验证输出结果...")
        output_errors = verify_output(config, source_headers)
        if output_errors:
            error_msg = "Output verification errors:\n" + "\n".join(f"  - {e}" for e in output_errors)
            fatal(error_msg)
        
    except InvalidFileException:
        traceback.print_exc()
        fatal(f"Error: Invalid Excel file '{input_path}'")
    except Exception as e:
        traceback.print_exc()
        fatal(f"Error processing Excel file: {e}")

def fatal(message):
    print(message)
    try:
        input("执行失败，按回车键退出")
    except EOFError:
        pass
    sys.exit(1)

def main():
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Process Excel file according to configuration')
    parser.add_argument('--config', default='config.yaml', help='Path to configuration file')
    args = parser.parse_args()

    # Load configuration
    config = load_config(args.config)
    
    # Process Excel file
    process_excel(config)

    input("执行成功，按回车键退出")


if __name__ == '__main__':
    main() 