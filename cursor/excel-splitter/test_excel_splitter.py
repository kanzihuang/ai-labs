import unittest
import os
import sys
import yaml
from io import StringIO
from openpyxl import Workbook, load_workbook
from main import process_excel, load_config

class TestExcelSplitter(unittest.TestCase):
    def setUp(self):
        # Create test configuration
        self.config = {
            'input': {
                'path': 'test_input.xlsx',
                'sheet': {
                    'source': {
                        'name': '工资',
                        'columns': {
                            'employee_id': '工号',
                            'project_id': '费用所属中心',
                            'project_category': '费用类别',
                            'project_hours': '实际出勤',
                            'employer_name': '工资所属单位'
                        }
                    },
                    'reference': {
                        'name': '工时',
                        'columns': {
                            'employee_id': '工号',
                            'project_id': '费用所属中心',
                            'project_category': '费用类别',
                            'project_hours': '实际出勤'
                        }
                    },
                    'payment': {
                        'name': '支付规则',
                        'columns': {
                            'project_id': '费用所属中心',
                            'employer_name': '公司',
                            'project_category': '费用类别'
                        },
                        'output_columns': ['支付账号']
                    }
                },
                'splitting_columns': ['基本工资', '岗位工资']
            },
            'output': {
                'path': 'test_output.xlsx',
                'sheet': {
                    'result': {
                        'name': '工资拆分'
                    }
                }
            }
        }

        # Create test input Excel file
        self.wb = Workbook()
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        # Create our sheets
        self.source_sheet = self.wb.create_sheet('工资')
        self.reference_sheet = self.wb.create_sheet('工时')
        self.payment_sheet = self.wb.create_sheet('支付规则')

    def tearDown(self):
        # Clean up test files
        if os.path.exists('test_input.xlsx'):
            os.remove('test_input.xlsx')
        if os.path.exists('test_output.xlsx'):
            os.remove('test_output.xlsx')

    def test_headers_consistency(self):
        """Test case for ensuring headers consistency between source and result"""
        # Set up source sheet headers according to test case 1
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)

        # Set up reference sheet headers
        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)

        # Set up payment sheet headers
        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)

        # Save the workbook
        self.wb.save('test_input.xlsx')

        # Process the Excel file
        process_excel(self.config)

        # Load output file and check headers
        output_wb = load_workbook('test_output.xlsx')
        result_sheet = output_wb['工资拆分']

        # Result headers should be source headers + payment_account column
        result_headers = [cell.value for cell in result_sheet[1]]
        expected_headers = source_headers + ['支付账号']
        self.assertEqual(expected_headers, result_headers)

    def test_direct_copy(self):
        """Test case for copying rows when no reference match exists"""
        # Set up source sheet according to test case 2
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)

        # Add data row to source
        source_data = ['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A']
        self.source_sheet.append(source_data)

        # Set up reference sheet without matching employee ID
        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        # No data in reference that matches the source

        # Set up payment sheet with mapping for source's project_id
        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        # Save the workbook
        self.wb.save('test_input.xlsx')

        # Process the Excel file
        process_excel(self.config)

        # Load output file and check data
        output_wb = load_workbook('test_output.xlsx')
        result_sheet = output_wb['工资拆分']

        # Check if data in row 2 matches the source data plus payment_account
        result_data = [cell.value for cell in result_sheet[2]]
        expected_data = source_data + ['Account_RD']
        self.assertEqual(expected_data, result_data)

    def test_splitting(self):
        """Test case for splitting rows based on reference data"""
        # Set up source sheet according to test case 3
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)

        # Add data row to source
        source_data = ['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A']
        self.source_sheet.append(source_data)

        # Set up reference sheet with matching employee ID but different project allocations
        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)

        # Add reference data rows
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 1])
        self.reference_sheet.append(['张三', 'AA', '研发', '2', 4])
        self.reference_sheet.append(['张三', 'AA', '销售', '3', 4])

        # Set up payment sheet: each project maps to a distinct account (no merging)
        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account1'])
        self.payment_sheet.append(['2', '公司A', '研发', 'Account2'])
        self.payment_sheet.append(['3', '公司A', '销售', 'Account3'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        # Save the workbook
        self.wb.save('test_input.xlsx')

        # Process the Excel file
        process_excel(self.config)

        # Load output file and check data
        output_wb = load_workbook('test_output.xlsx')
        result_sheet = output_wb['工资拆分']

        # Expected results with payment_account column added (no merge since accounts differ)
        expected_results = [
            ['张三', 'AA', '中国', 111.11, 333.33, '研发', '1', 1, 'BB', '公司A', 'Account1'],
            ['张三', 'AA', '中国', 444.44, 1333.33, '研发', '2', 4, 'BB', '公司A', 'Account2'],
            ['张三', 'AA', '中国', 444.44, 1333.33, '销售', '3', 4, 'BB', '公司A', 'Account3']
        ]

        # Check if each row matches the expected split results
        for i, expected in enumerate(expected_results, start=2):
            result_row = [cell.value for cell in result_sheet[i]]

            # Compare each value with a tolerance for floating point values
            for j, (expected_val, actual_val) in enumerate(zip(expected, result_row)):
                if isinstance(expected_val, (int, float)) and isinstance(actual_val, (int, float)):
                    self.assertAlmostEqual(expected_val, actual_val, places=1)
                else:
                    self.assertEqual(expected_val, actual_val)

    # --- New tests for merge by payment account ---
    def test_reference_duplicate_employee_project(self):
        """Reference table has duplicate (employee_id, project_id) -> fatal"""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        # Duplicate (AA, 1) pairs
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 1])
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 4])

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_X'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        with self.assertRaises(SystemExit):
            process_excel(self.config)

    def test_payment_duplicate_project_id(self):
        """Payment table has duplicate project_id -> fatal"""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 1])

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_X'])
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_Y'])  # Duplicate (employer, project_id, category)
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        with self.assertRaises(SystemExit):
            process_excel(self.config)

    def test_payment_empty_account(self):
        """Payment has non-empty project_id with empty payment_account -> allowed (no check)"""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 1])

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', None])  # Empty account — allowed
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        # Should NOT raise SystemExit — empty output columns are allowed
        process_excel(self.config)

    def test_source_project_not_in_payment(self):
        """Source project_id missing from payment -> fatal"""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        # Source project_id = "研发部" but no payment mapping for it
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        # Only project "1", no "研发部"
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_X'])

        self.wb.save('test_input.xlsx')
        with self.assertRaises(SystemExit):
            process_excel(self.config)

    def test_reference_project_not_in_payment(self):
        """Reference has project_id not in payment -> fatal"""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        # Reference has project_id "99" not in payment
        self.reference_sheet.append(['张三', 'AA', '研发', '99', 5])

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_X'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        with self.assertRaises(SystemExit):
            process_excel(self.config)

    def test_source_project_not_in_payment_but_will_be_split(self):
        """Source project_id not in payment is OK if row will be split (project_id comes from reference)"""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        # Source has project_id "研发部" not in payment, but employee has reference match
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 1])
        self.reference_sheet.append(['张三', 'AA', '研发', '2', 4])

        # Payment has reference project_ids but NOT source's "研发部"
        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_X'])
        self.payment_sheet.append(['2', '公司A', '研发', 'Account_X'])

        self.wb.save('test_input.xlsx')
        # Should NOT fatal — source's project_id "研发部" will be replaced by reference's project_ids
        # which are both in payment
        process_excel(self.config)

        # Verify output exists and has 2 rows (no merge)
        output_wb = load_workbook('test_output.xlsx')
        result_sheet = output_wb['工资拆分']
        result_rows = list(result_sheet.iter_rows(min_row=2))
        self.assertEqual(len(result_rows), 2)
        self.assertEqual(result_rows[0][6].value, '1')
        self.assertEqual(result_rows[0][10].value, 'Account_X')
        self.assertEqual(result_rows[1][6].value, '2')
        self.assertEqual(result_rows[1][10].value, 'Account_X')


    # --- Merge key correctness ---
    def test_mixed_split_and_no_split_rows(self):
        """One source row splits, another doesn't — both handled correctly"""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        # Row 1: has reference match → will be split, source project_id ignored
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', 'no_such_project', 21, 'BB', '公司A'])
        # Row 2: no reference match → copied as-is, source project_id MUST be in payment
        self.source_sheet.append(['李四', 'CC', '中国', 2000.00, 5000.00, '研发', '研发部', 21, 'DD', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 5])
        self.reference_sheet.append(['张三', 'AA', '研发', '2', 5])

        # Payment has reference project_ids AND source row 2's project_id, but NOT row 1's
        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_X'])
        self.payment_sheet.append(['2', '公司A', '研发', 'Account_Y'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])
        # Note: no "no_such_project" in payment, but that's OK because row 1 splits

        self.wb.save('test_input.xlsx')
        process_excel(self.config)

        output_wb = load_workbook('test_output.xlsx')
        result_sheet = output_wb['工资拆分']

        result_rows = list(result_sheet.iter_rows(min_row=2))
        self.assertEqual(len(result_rows), 3,
                         "Row 1 splits into 2 (distinct accounts) + Row 2 copied as 1 = 3 rows")

        # Row 1 split results: project "1" -> Account_X, project "2" -> Account_Y
        row1_data = [cell.value for cell in result_rows[0]]
        row2_data = [cell.value for cell in result_rows[1]]
        self.assertIn(row1_data[6], ['1', '2'])
        self.assertIn(row1_data[10], ['Account_X', 'Account_Y'])

        # Row 2 (no split): should retain project_id "研发部" and get Account_RD
        row3_data = [cell.value for cell in result_rows[2]]
        self.assertEqual(row3_data[1], 'CC')
        self.assertEqual(row3_data[6], '研发部')
        self.assertEqual(row3_data[10], 'Account_RD')

    # --- Config validation tests ---

    def _make_minimal_config(self):
        """Helper: minimal valid config dict."""
        return {
            'input': {
                'path': 'test_input.xlsx',
                'sheet': {
                    'source': {
                        'name': '工资',
                        'columns': {
                            'employee_id': '工号',
                            'project_id': '费用所属中心',
                            'project_category': '费用类别',
                            'project_hours': '实际出勤',
                            'employer_name': '工资所属单位'
                        }
                    },
                    'reference': {
                        'name': '工时',
                        'columns': {
                            'employee_id': '工号',
                            'project_id': '费用所属中心',
                            'project_category': '费用类别',
                            'project_hours': '实际出勤'
                        }
                    },
                    'payment': {
                        'name': '支付规则',
                        'columns': {
                            'project_id': '费用所属中心',
                            'employer_name': '公司',
                            'project_category': '费用类别'
                        },
                        'output_columns': ['支付账号']
                    }
                },
                'splitting_columns': ['基本工资']
            },
            'output': {
                'path': 'test_output.xlsx',
                'sheet': {
                    'result': {
                        'name': '工资拆分'
                    }
                }
            }
        }

    def test_valid_config_passes_check(self):
        """Valid config should pass validate_config without error."""
        from main import validate_config
        config = self._make_minimal_config()
        cc = validate_config(config)
        self.assertEqual(cc, {})  # no computed_columns by default

    def test_config_missing_input_section(self):
        """Config missing 'input' key -> SystemExit."""
        from main import validate_config
        with self.assertRaises(SystemExit):
            validate_config({'output': {}})

    def test_config_missing_input_path(self):
        """Config missing input.path -> SystemExit."""
        from main import validate_config
        config = self._make_minimal_config()
        del config['input']['path']
        with self.assertRaises(SystemExit):
            validate_config(config)

    def test_config_missing_splitting_columns(self):
        """Config missing splitting_columns -> SystemExit."""
        from main import validate_config
        config = self._make_minimal_config()
        del config['input']['splitting_columns']
        with self.assertRaises(SystemExit):
            validate_config(config)

    def test_config_non_list_splitting_columns(self):
        """Config with non-list splitting_columns -> SystemExit."""
        from main import validate_config
        config = self._make_minimal_config()
        config['input']['splitting_columns'] = 'not_a_list'
        with self.assertRaises(SystemExit):
            validate_config(config)

    def test_config_missing_output_section(self):
        """Config missing 'output' section -> SystemExit."""
        from main import validate_config
        config = self._make_minimal_config()
        del config['output']
        with self.assertRaises(SystemExit):
            validate_config(config)

    def test_config_missing_source_name(self):
        """Config missing source.name -> SystemExit."""
        from main import validate_config
        config = self._make_minimal_config()
        del config['input']['sheet']['source']['name']
        with self.assertRaises(SystemExit):
            validate_config(config)

    def test_config_missing_reference_columns(self):
        """Config missing reference columns -> SystemExit."""
        from main import validate_config
        config = self._make_minimal_config()
        del config['input']['sheet']['reference']['columns']['project_hours']
        with self.assertRaises(SystemExit):
            validate_config(config)

    # --- Data quality pre-check tests ---

    def test_none_employee_id_in_reference(self):
        """Reference row with None employee_id -> SystemExit."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', None, '研发', '1', 5])  # None employee_id

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_X'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        with self.assertRaises(SystemExit):
            process_excel(self.config)

    def test_none_employee_id_in_source(self):
        """Source row with None employee_id -> SystemExit."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', None, '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 5])

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_X'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        with self.assertRaises(SystemExit):
            process_excel(self.config)

    def test_negative_hours_in_reference(self):
        """Reference row with negative hours -> SystemExit."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', -5])  # Negative hours

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_X'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        with self.assertRaises(SystemExit):
            process_excel(self.config)

    def test_non_numeric_hours_in_reference(self):
        """Reference row with text in hours column -> SystemExit."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 'N/A'])  # Non-numeric hours

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_X'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        with self.assertRaises(SystemExit):
            process_excel(self.config)

    def test_employee_id_type_mismatch(self):
        """Mixed int/str employee_id across sheets -> SystemExit."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        # Reference uses int type for employee_id, source uses str
        self.reference_sheet.append(['张三', 123, '研发', '1', 5])

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_X'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        with self.assertRaises(SystemExit):
            process_excel(self.config)

    # --- Output verification tests ---

    def test_output_sum_consistency(self):
        """Splitting column totals in output should match source sums."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A'])
        self.source_sheet.append(['李四', 'BB', '中国', 2000.00, 4000.00, '研发', '研发部', 21, 'CC', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 2])
        self.reference_sheet.append(['张三', 'AA', '研发', '2', 3])

        # Distinct accounts
        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_1'])
        self.payment_sheet.append(['2', '公司A', '研发', 'Account_2'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        process_excel(self.config)

        # Verify output sums
        output_wb = load_workbook('test_output.xlsx')
        result_sheet = output_wb['工资拆分']

        for col_idx in [4, 5]:  # 基本工资 and 岗位工资
            result_sum = sum(
                float(row[col_idx - 1].value)
                for row in result_sheet.iter_rows(min_row=2)
                if row[col_idx - 1].value is not None
            )
            source_sum = sum(
                float(row[col_idx - 1].value)
                for row in self.source_sheet.iter_rows(min_row=2)
                if row[col_idx - 1].value is not None
            )
            self.assertAlmostEqual(result_sum, source_sum, places=1)

    def test_output_file_valid(self):
        """Output file can be reopened and has expected structure."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '1', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 5])

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_X'])

        self.wb.save('test_input.xlsx')
        process_excel(self.config)

        # Verify output file
        self.assertTrue(os.path.exists('test_output.xlsx'))
        wb = load_workbook('test_output.xlsx')
        self.assertIn('工资拆分', wb.sheetnames)
        self.assertIn('工资', wb.sheetnames)
        self.assertIn('工时', wb.sheetnames)
        wb.close()

    # --- 3-field payment matching tests ---
    def test_composite_key_same_project_diff_employer_ok(self):
        """Same project_id with different employer_name is NOT a duplicate — both valid."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 10])

        # Same project_id "1" with two different employers — both valid
        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_X'])
        self.payment_sheet.append(['1', '公司B', '研发', 'Account_Y'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_X'])

        self.wb.save('test_input.xlsx')
        process_excel(self.config)

        output_wb = load_workbook('test_output.xlsx')
        result_sheet = output_wb['工资拆分']
        result_row = [cell.value for cell in result_sheet[2]]
        # AA/公司A + project "1" → Account_X
        self.assertEqual(result_row[10], 'Account_X')

    def test_composite_key_different_employer_diff_account(self):
        """Same project_id + different employer_name → different accounts, not merged."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        # Same employee "AA", two source rows with different employers
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A'])
        self.source_sheet.append(['张三', 'AA', '中国', 500.00,  1500.00, '研发', '研发部', 21, 'BB', '公司B'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 10])

        # Same project_id "1", different employers → different accounts
        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_X'])
        self.payment_sheet.append(['1', '公司B', '研发', 'Account_Y'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_X'])
        self.payment_sheet.append(['研发部', '公司B', '研发', 'Account_Y'])

        self.wb.save('test_input.xlsx')
        process_excel(self.config)

        output_wb = load_workbook('test_output.xlsx')
        result_sheet = output_wb['工资拆分']
        result_rows = list(result_sheet.iter_rows(min_row=2))

        # Two rows (different source rows → not merged), different accounts
        self.assertEqual(len(result_rows), 2)
        row1 = [cell.value for cell in result_rows[0]]
        row2 = [cell.value for cell in result_rows[1]]
        self.assertEqual(row1[10], 'Account_X')
        self.assertEqual(row2[10], 'Account_Y')
        self.assertAlmostEqual(row1[3], 1000.00, places=1)
        self.assertAlmostEqual(row2[3], 500.00, places=1)

    def test_composite_key_missing_pair_in_payment(self):
        """Split row's (employer_name, project_id) not in payment → fatal."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        # Reference project_id "99" not in payment for 公司A
        self.reference_sheet.append(['张三', 'AA', '研发', '99', 10])

        # Payment has project "99" but only for 公司B — not 公司A
        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['99', '公司B', '研发', 'Account_Z'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        with self.assertRaises(SystemExit):
            process_excel(self.config)

    def test_composite_key_none_employer_in_payment(self):
        """Payment row with empty employer_name is skipped."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 10])

        # Payment has a row with empty employer_name (should be skipped)
        # and another row with valid employer_name for the same project_id
        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_X'])
        self.payment_sheet.append(['1', None, '研发', 'Account_IGNORED'])  # Empty employer → skipped
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        process_excel(self.config)

        output_wb = load_workbook('test_output.xlsx')
        result_sheet = output_wb['工资拆分']
        result_row = [cell.value for cell in result_sheet[2]]
        # Should match (公司A, 1) → Account_X, ignore the None-employer row
        self.assertEqual(result_row[10], 'Account_X')

    # --- Computed columns tests ---

    def _make_computed_config(self):
        """Helper: config with computed_columns."""
        return {
            'input': {
                'path': 'test_input.xlsx',
                'sheet': {
                    'source': {
                        'name': '工资',
                        'columns': {
                            'employee_id': '工号',
                            'project_id': '费用所属中心',
                            'project_category': '费用类别',
                            'project_hours': '实际出勤',
                            'employer_name': '工资所属单位'
                        }
                    },
                    'reference': {
                        'name': '工时',
                        'columns': {
                            'employee_id': '工号',
                            'project_id': '费用所属中心',
                            'project_category': '费用类别',
                            'project_hours': '实际出勤'
                        }
                    },
                    'payment': {
                        'name': '支付规则',
                        'columns': {
                            'project_id': '费用所属中心',
                            'employer_name': '公司',
                            'project_category': '费用类别'
                        },
                        'output_columns': ['支付账号']
                    }
                },
                'splitting_columns': ['基本工资', '岗位工资'],
                'computed_columns': {
                    '福利前工资合计': '基本工资 + 岗位工资'
                }
            },
            'output': {
                'path': 'test_output.xlsx',
                'sheet': {
                    'result': {
                        'name': '工资拆分'
                    }
                }
            }
        }

    def test_computed_columns_basic(self):
        """Computed column = sum of split base columns per result row."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '福利前工资合计',
                          '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, 4000.00,
                                   '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 1])
        self.reference_sheet.append(['张三', 'AA', '研发', '2', 4])

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_1'])
        self.payment_sheet.append(['2', '公司A', '研发', 'Account_2'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        config = self._make_computed_config()
        process_excel(config)

        output_wb = load_workbook('test_output.xlsx')
        result_sheet = output_wb['工资拆分']
        result_rows = list(result_sheet.iter_rows(min_row=2))

        # 2 split rows: project 1 (1/5) and project 2 (4/5)
        self.assertEqual(len(result_rows), 2)
        row1 = [cell.value for cell in result_rows[0]]
        row2 = [cell.value for cell in result_rows[1]]

        # Row 1: 基本工资=200, 岗位工资=600 → 福利前 = 800
        self.assertAlmostEqual(row1[3], 200.00, places=1)
        self.assertAlmostEqual(row1[4], 600.00, places=1)
        self.assertAlmostEqual(row1[5], 800.00, places=1)

        # Row 2: 基本工资=800, 岗位工资=2400 → 福利前 = 3200
        self.assertAlmostEqual(row2[3], 800.00, places=1)
        self.assertAlmostEqual(row2[4], 2400.00, places=1)
        self.assertAlmostEqual(row2[5], 3200.00, places=1)

    def test_computed_columns_chained(self):
        """Three chained formulas compute correctly."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '福利前工资合计',
                          '用餐补助', '税前应发工资总额(不含差补)', '差旅补助', '实发工资',
                          '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, 4000.00,
                                   500.00, 4500.00, 200.00, 4700.00,
                                   '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 5])

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_1'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        config = {
            'input': {
                'path': 'test_input.xlsx',
                'sheet': {
                    'source': {
                        'name': '工资',
                        'columns': {
                            'employee_id': '工号',
                            'project_id': '费用所属中心',
                            'project_category': '费用类别',
                            'project_hours': '实际出勤',
                            'employer_name': '工资所属单位'
                        }
                    },
                    'reference': {
                        'name': '工时',
                        'columns': {
                            'employee_id': '工号',
                            'project_id': '费用所属中心',
                            'project_category': '费用类别',
                            'project_hours': '实际出勤'
                        }
                    },
                    'payment': {
                        'name': '支付规则',
                        'columns': {
                            'project_id': '费用所属中心',
                            'employer_name': '公司',
                            'project_category': '费用类别'
                        },
                        'output_columns': ['支付账号']
                    }
                },
                'splitting_columns': ['基本工资', '岗位工资', '用餐补助', '差旅补助'],
                'computed_columns': {
                    '福利前工资合计': '基本工资 + 岗位工资',
                    '税前应发工资总额(不含差补)': '福利前工资合计 + 用餐补助',
                    '实发工资': '税前应发工资总额(不含差补) + 差旅补助'
                }
            },
            'output': {
                'path': 'test_output.xlsx',
                'sheet': {
                    'result': {
                        'name': '工资拆分'
                    }
                }
            }
        }
        process_excel(config)

        output_wb = load_workbook('test_output.xlsx')
        result_sheet = output_wb['工资拆分']
        result_rows = list(result_sheet.iter_rows(min_row=2))

        self.assertEqual(len(result_rows), 1)
        row = [cell.value for cell in result_rows[0]]

        # All amounts go to single project (ratio=1.0)
        self.assertAlmostEqual(row[3], 1000.00, places=1)  # 基本工资
        self.assertAlmostEqual(row[4], 3000.00, places=1)  # 岗位工资
        self.assertAlmostEqual(row[5], 4000.00, places=1)  # 福利前 = 1000+3000
        self.assertAlmostEqual(row[6], 500.00, places=1)   # 用餐补助
        self.assertAlmostEqual(row[7], 4500.00, places=1)  # 税前 = 4000+500
        self.assertAlmostEqual(row[8], 200.00, places=1)   # 差旅补助
        self.assertAlmostEqual(row[9], 4700.00, places=1)  # 实发 = 4500+200
    def test_computed_unsplit_rows_unchanged(self):
        """Non-split rows keep computed column values unchanged."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '福利前工资合计',
                          '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        # Source has a pre-existing 福利前=9999 (not matching 基本+岗位=4000)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, 9999.00,
                                   '研发', '研发部', 21, 'BB', '公司A'])

        # No matching reference → row not split
        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        config = self._make_computed_config()
        process_excel(config)

        output_wb = load_workbook('test_output.xlsx')
        result_sheet = output_wb['工资拆分']
        result_rows = list(result_sheet.iter_rows(min_row=2))

        self.assertEqual(len(result_rows), 1)
        row = [cell.value for cell in result_rows[0]]
        # Unsplit row: computed column keeps original source value (9999), not recomputed
        self.assertEqual(row[5], 9999.00)

    def test_computed_column_not_in_source_fatal(self):
        """Computed column not in source headers → fatal."""
        # Source has no 福利前工资合计 column
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资',
                          '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00,
                                   '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 5])

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_1'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        config = self._make_computed_config()
        with self.assertRaises(SystemExit):
            process_excel(config)

    # --- Computed column validation error tests ---

    def test_computed_in_splitting_fatal(self):
        """Computed column name in splitting_columns → fatal."""
        from main import validate_config
        config = self._make_computed_config()
        config['input']['splitting_columns'].append('福利前工资合计')
        with self.assertRaises(SystemExit):
            validate_config(config)

    def test_computed_unknown_ref_fatal(self):
        """Formula references nonexistent column → fatal in validate_sheets."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '福利前工资合计',
                          '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, 4000.00,
                                   '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 5])

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_1'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        config = self._make_computed_config()
        # Formula references '不存在的列' which doesn't exist
        config['input']['computed_columns']['福利前工资合计'] = '基本工资 + 不存在的列'
        with self.assertRaises(SystemExit):
            process_excel(config)

    def test_computed_ref_not_in_splitting_fatal(self):
        """Formula references a column not in splitting_columns → fatal."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '福利前工资合计',
                          '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, 4000.00,
                                   '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 5])

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_1'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        config = self._make_computed_config()
        # '部门' exists in source but NOT in splitting_columns
        config['input']['computed_columns']['福利前工资合计'] = '基本工资 + 部门'
        with self.assertRaises(SystemExit):
            process_excel(config)

    def test_computed_circular_fatal(self):
        """Formula references a later computed column → fatal."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '福利前工资合计',
                          '用餐补助', '税前应发工资总额(不含差补)',
                          '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, 4000.00,
                                   500.00, 4500.00,
                                   '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 5])

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_1'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        config = {
            'input': {
                'path': 'test_input.xlsx',
                'sheet': {
                    'source': {
                        'name': '工资',
                        'columns': {
                            'employee_id': '工号',
                            'project_id': '费用所属中心',
                            'project_category': '费用类别',
                            'project_hours': '实际出勤',
                            'employer_name': '工资所属单位'
                        }
                    },
                    'reference': {
                        'name': '工时',
                        'columns': {
                            'employee_id': '工号',
                            'project_id': '费用所属中心',
                            'project_category': '费用类别',
                            'project_hours': '实际出勤'
                        }
                    },
                    'payment': {
                        'name': '支付规则',
                        'columns': {
                            'project_id': '费用所属中心',
                            'employer_name': '公司',
                            'project_category': '费用类别'
                        },
                        'output_columns': ['支付账号']
                    }
                },
                'splitting_columns': ['基本工资', '岗位工资', '用餐补助'],
                'computed_columns': {
                    # 福利前 references 税前 which is defined later → error
                    '福利前工资合计': '基本工资 + 税前应发工资总额(不含差补)',
                    '税前应发工资总额(不含差补)': '福利前工资合计 + 用餐补助'
                }
            },
            'output': {
                'path': 'test_output.xlsx',
                'sheet': {
                    'result': {
                        'name': '工资拆分'
                    }
                }
            }
        }
        with self.assertRaises(SystemExit):
            process_excel(config)

    def test_verify_computed_source_row_mismatch(self):
        """Per-source-row sum mismatch is fatal."""
        source_headers = ['姓名', '工号', '部门', '基本工资', '岗位工资', '福利前工资合计',
                          '费用类别', '费用所属中心', '实际出勤', '分管领导', '工资所属单位']
        self.source_sheet.append(source_headers)
        self.source_sheet.append(['张三', 'AA', '中国', 1000.00, 3000.00, 4000.00,
                                   '研发', '研发部', 21, 'BB', '公司A'])

        reference_headers = ['姓名', '工号', '费用类别', '费用所属中心', '实际出勤']
        self.reference_sheet.append(reference_headers)
        self.reference_sheet.append(['张三', 'AA', '研发', '1', 1])
        self.reference_sheet.append(['张三', 'AA', '研发', '2', 4])

        payment_headers = ['费用所属中心', '公司', '费用类别', '支付账号']
        self.payment_sheet.append(payment_headers)
        self.payment_sheet.append(['1', '公司A', '研发', 'Account_1'])
        self.payment_sheet.append(['2', '公司A', '研发', 'Account_2'])
        self.payment_sheet.append(['研发部', '公司A', '研发', 'Account_RD'])

        self.wb.save('test_input.xlsx')
        # Re-open and modify the saved file so source value doesn't match computed sum
        wb2 = load_workbook('test_input.xlsx')
        wb2['工资']['F2'] = 5000.00
        wb2.save('test_input.xlsx')
        wb2.close()
        config = self._make_computed_config()
        with self.assertRaises(SystemExit):
            process_excel(config)

    # --- Formula parser tests ---

    def test_formula_parser_basic(self):
        """Basic formula parsing and evaluation."""
        from main import tokenize_formula, parse_formula, evaluate_formula
        known = ['基本工资', '岗位工资']
        tokens = tokenize_formula('基本工资 + 岗位工资', known)
        ast = parse_formula(tokens)
        result = evaluate_formula(ast, {'基本工资': 100.0, '岗位工资': 200.0})
        self.assertEqual(result, 300.0)

    def test_formula_parser_precedence(self):
        """Operator precedence: * before +."""
        from main import tokenize_formula, parse_formula, evaluate_formula
        known = ['基本工资', '岗位工资']
        tokens = tokenize_formula('基本工资 + 岗位工资 * 2', known)
        ast = parse_formula(tokens)
        result = evaluate_formula(ast, {'基本工资': 100.0, '岗位工资': 200.0})
        self.assertEqual(result, 500.0)  # 100 + (200*2) = 500

    def test_formula_parser_parens(self):
        """Parentheses override precedence."""
        from main import tokenize_formula, parse_formula, evaluate_formula
        known = ['基本工资', '岗位工资']
        tokens = tokenize_formula('(基本工资 + 岗位工资) * 2', known)
        ast = parse_formula(tokens)
        result = evaluate_formula(ast, {'基本工资': 100.0, '岗位工资': 200.0})
        self.assertEqual(result, 600.0)  # (100+200)*2 = 600

    def test_formula_parser_numeric_literal(self):
        """Numeric literals in formulas."""
        from main import tokenize_formula, parse_formula, evaluate_formula
        known = ['基本工资']
        tokens = tokenize_formula('基本工资 * 1.5', known)
        ast = parse_formula(tokens)
        result = evaluate_formula(ast, {'基本工资': 100.0})
        self.assertEqual(result, 150.0)

    def test_formula_parser_name_with_parens(self):
        """Column name containing parentheses is matched as one token."""
        from main import tokenize_formula, parse_formula, evaluate_formula
        known = ['税前应发工资总额(不含差补)', '差旅补助']
        tokens = tokenize_formula('税前应发工资总额(不含差补) + 差旅补助', known)
        ast = parse_formula(tokens)
        result = evaluate_formula(ast, {'税前应发工资总额(不含差补)': 4500.0, '差旅补助': 200.0})
        self.assertEqual(result, 4700.0)

    def test_formula_parser_division_by_zero(self):
        """Division by zero raises FormulaError."""
        from main import tokenize_formula, parse_formula, evaluate_formula, FormulaError
        known = ['基本工资']
        tokens = tokenize_formula('基本工资 / 0', known)
        ast = parse_formula(tokens)
        with self.assertRaises(FormulaError):
            evaluate_formula(ast, {'基本工资': 100.0})

    def test_formula_parser_malformed(self):
        """Malformed formula raises FormulaError."""
        from main import tokenize_formula, parse_formula, FormulaError
        known = ['基本工资']
        tokens = tokenize_formula('基本工资 +', known)
        with self.assertRaises(FormulaError):
            parse_formula(tokens)

    def test_formula_parser_none_value(self):
        """None values are treated as 0.0."""
        from main import tokenize_formula, parse_formula, evaluate_formula
        known = ['基本工资', '岗位工资']
        tokens = tokenize_formula('基本工资 + 岗位工资', known)
        ast = parse_formula(tokens)
        result = evaluate_formula(ast, {'基本工资': 100.0, '岗位工资': None})
        self.assertEqual(result, 100.0)


if __name__ == '__main__':
    unittest.main()
